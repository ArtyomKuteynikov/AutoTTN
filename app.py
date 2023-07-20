import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
import requests
from openpyxl import load_workbook
from openpyxl.styles import Border, Side


def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)
    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment
    rows = ws[cell_range]
    if font:
        first_cell.font = font
    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom
    for row in rows:
        l = row[0]
        r = row[-1]
        # l.border = l.border + left
        # r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill


def get_company_info(company_name):
    api_key = "63b6e9bd311c26253227a11acfb6397229245935"  # Замените на ваш ключ API DaData
    api_url = "https://suggestions.dadata.ru/suggestions/api/4_1/rs/suggest/party"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Token {api_key}"
    }
    payload = {
        "query": company_name,
        "count": 1
    }
    try:
        response = requests.post(api_url, json=payload, headers=headers)
        data = response.json()
        if response.status_code == 200:
            suggestions = data.get("suggestions", [])
            if suggestions:
                suggestion = suggestions[0]
                data = suggestion.get("data", {})
                return f'{data.get("name").get("short_with_opf")}, ИНН: {data.get("inn")}, {data.get("address").get("value")}'
            else:
                return "Компания не найдена."
        else:
            return "Ошибка при выполнении запроса к API DaData."
    except requests.exceptions.RequestException as e:
        return "Ошибка при подключении к API DaData:", str(e)


def get_address(company_name):
    api_key = "63b6e9bd311c26253227a11acfb6397229245935"  # Замените на ваш ключ API DaData
    api_url = "https://suggestions.dadata.ru/suggestions/api/4_1/rs/suggest/party"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Token {api_key}"
    }
    payload = {
        "query": company_name,
        "count": 1
    }
    try:
        response = requests.post(api_url, json=payload, headers=headers)
        data = response.json()
        if response.status_code == 200:
            suggestions = data.get("suggestions", [])
            if suggestions:
                suggestion = suggestions[0]
                data = suggestion.get("data", {})
                return f'{data.get("address").get("value")}'
            else:
                return "Компания не найдена."
        else:
            return "Ошибка при выполнении запроса к API DaData."
    except requests.exceptions.RequestException as e:
        return "Ошибка при подключении к API DaData:", str(e)



def clear():
    sender_combobox.delete("0", tk.END)
    sender_entry.delete("0", tk.END)
    documents_entry.delete("0", tk.END)
    recipient_entry.delete("0", tk.END)
    cargo_text.delete("1.0","end")
    carrier_entry.delete("0", tk.END)
    driver_entry.delete("0", tk.END)
    model_entry.delete("0", tk.END)
    license_plate_entry.delete("0", tk.END)


def validate_input():
    try:
        sender = sender_var.get()
        sender_fio = sender_fio_var.get()
        cargo = cargo_text.get("1.0", "end-1c")
        documents = documents_var.get()
        recipient = recipient_var.get()
        carrier = carrier_var.get()
        driver = driver_var.get()
        model = model_var.get()
        license_plate = license_plate_var.get()
        if not sender or not cargo or not documents or not recipient or not carrier or not driver or not model or not license_plate:
            messagebox.showerror("Ошибка", "Пожалуйста, заполните все поля.")
            return
        file = r"Транспортная накладная ШАБЛОН.xlsx"
        try:
            wb = load_workbook(file)
        except:
            messagebox.showerror("Ошибка", f"Не найден файл шаблона, куда дели, блин его???")
            return
        ws = wb.active
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for range in ws.merged_cells.ranges:
            style_range(ws, str(range), border=border)
        # Получатель
        ws['AD13'] = get_company_info(recipient)
        ws['AD47'] = get_company_info(recipient)
        ws['AD49'] = get_address(recipient)
        # Перевозчик
        ws['AD124'] = get_company_info(carrier)
        ws['B82'] = get_company_info(carrier)
        # Водитель
        ws['AD61'] = driver
        ws['AD114'] = driver
        ws['AD126'] = driver
        ws['B61'] = driver
        ws['B86'] = driver
        # Транспортное средство
        ws['B91'] = model
        ws['AD93'] = license_plate
        # Груз, сопр документы
        ws['B21'] = cargo
        ws['B32'] = documents
        # Отправитель
        ws['B59'] = sender_fio
        ws['B114'] = sender_fio
        ws['B126'] = sender_fio
        ws['B13'] = get_company_info(sender)
        ws['B47'] = get_company_info(sender)
        ws['B124'] = get_company_info(sender)
        # Дата
        ws['D8'] = get_current_date()
        ws['AF8'] = get_current_date()
        ws['B51'] = get_current_date()
        ws['B53'] = get_current_date()
        ws['A77'] = get_current_date()
        ws['B128'] = get_current_date()
        ws['AD128'] = get_current_date()
        try:
            wb.save(f"Транспортная накладная от {get_current_date()} {recipient}.xlsx")
        except:
            messagebox.showerror("Ошибка", f"Закройте файл {f'Транспортная накладная от {get_current_date()} {recipient}.xlsx'}")
            return
        wb.close()
        messagebox.showinfo("Успешно", "ТТН успешно создана!")
    except Exception as e:
        messagebox.showerror("Ошибка", str(e))
        return


def get_current_date():
    import datetime
    now = datetime.datetime.now()
    return now.strftime("%d.%m.%Y")


root = tk.Tk()
root.title("Генератор ТТН")
root.geometry("400x600")

sender_var = tk.StringVar()
sender_fio_var = tk.StringVar()
cargo_text = tk.Text(root, height=5)
documents_var = tk.StringVar()
recipient_var = tk.StringVar()
carrier_var = tk.StringVar()
driver_var = tk.StringVar()
model_var = tk.StringVar()
license_plate_var = tk.StringVar()

sender_label = tk.Label(root, text="Отправитель:")
sender_label.pack()
sender_label = tk.Label(root, text="Отправитель(Юр. лицо):")
sender_label.pack()
sender_combobox = ttk.Combobox(root, textvariable=sender_var,
                               values=["АО СпецАвтоИнжиниринг", "ООО СпецАвтоИнжиниринг"])
sender_combobox.pack()
sender_label = tk.Label(root, text="Отправитель(ФИО):")
sender_label.pack()
sender_entry = tk.Entry(root, textvariable=sender_fio_var)
sender_entry.pack()

cargo_label = tk.Label(root, text="Груз:")
cargo_label.pack()
cargo_text.pack()

documents_label = tk.Label(root, text="Сопроводительные документы:")
documents_label.pack()
documents_entry = tk.Entry(root, textvariable=documents_var)
documents_entry.pack()

recipient_label = tk.Label(root, text="Получатель:")
recipient_label.pack()
recipient_entry = tk.Entry(root, textvariable=recipient_var)
recipient_entry.pack()

carrier_label = tk.Label(root, text="Перевозчик:")
carrier_label.pack()
carrier_entry = tk.Entry(root, textvariable=carrier_var)
carrier_entry.pack()

driver_label = tk.Label(root, text="Водитель:")
driver_label.pack()
driver_entry = tk.Entry(root, textvariable=driver_var)
driver_entry.pack()

vehicle_label = tk.Label(root, text="Транспортное средство:")
vehicle_label.pack()

model_label = tk.Label(root, text="  Модель, марка:")
model_label.pack()
model_entry = tk.Entry(root, textvariable=model_var)
model_entry.pack()

license_plate_label = tk.Label(root, text="  Госномер:")
license_plate_label.pack()
license_plate_entry = tk.Entry(root, textvariable=license_plate_var)
license_plate_entry.pack()

submit_button = tk.Button(root, text="Сформировать", command=validate_input)
submit_button.pack()
clear_button = tk.Button(root, text="Очистить", command=clear)
clear_button.pack()

license_plate_label = tk.Label(root, text="©Артём Кутейников")
license_plate_label.pack()

root.mainloop()
